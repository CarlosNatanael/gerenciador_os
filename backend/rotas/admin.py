from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
import io
from functools import wraps
from datetime import datetime
from werkzeug.security import generate_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from database import get_db

admin_bp = Blueprint('admin', __name__)

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('tipo') not in ['admin', 'master-admin']:
            flash('Acesso não autorizado.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated_function

@admin_bp.route('/')
@admin_required
def admin_dashboard():
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT 
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'Concluída' THEN 1 ELSE 0 END) AS concluidas,
                SUM(CASE WHEN status = 'Aberta' THEN 1 ELSE 0 END) AS abertas,
                SUM(CASE WHEN status = 'Em andamento' THEN 1 ELSE 0 END) AS em_andamento,
                SUM(CASE WHEN status = 'Agendada' THEN 1 ELSE 0 END) AS agendadas,
                AVG(CASE WHEN status = 'Concluída' AND tempo_reparo IS NOT NULL THEN tempo_reparo ELSE NULL END) AS media_tempo
            FROM ordens_servico
        """)
        stats_os_data = cursor.fetchone()

        stats_os_dict = dict(stats_os_data) if stats_os_data else {
            'total': 0, 'concluidas': 0, 'abertas': 0, 
            'em_andamento': 0, 'agendadas': 0, 'media_tempo': 0.0
        }
        if stats_os_dict['media_tempo'] is None: 
            stats_os_dict['media_tempo'] = 0.0

        cursor.execute("""
            SELECT 
                COUNT(*) AS total_registros,
                SUM(CASE WHEN status = 'Pendente Aprovacao' THEN 1 ELSE 0 END) AS pendente_aprovacao,
                SUM(CASE WHEN status = 'Concluido' THEN 1 ELSE 0 END) AS concluidos_registros,
                SUM(CASE WHEN status = 'Cancelado' THEN 1 ELSE 0 END) AS cancelados_registros
            FROM registros_manutencao_direta
        """)
        stats_registros_data = cursor.fetchone()
        
        stats_registros_dict = dict(stats_registros_data) if stats_registros_data else {
            'total_registros': 0, 'pendente_aprovacao': 0, 
            'concluidos_registros': 0, 'cancelados_registros': 0
        }
        stats_registros_dict.setdefault('total_registros', 0)
        stats_registros_dict.setdefault('pendente_aprovacao', 0)
        stats_registros_dict.setdefault('concluidos_registros', 0)
        stats_registros_dict.setdefault('cancelados_registros', 0)

        cursor.execute("""
            SELECT os.*, u.nome as solicitante_nome
            FROM ordens_servico os
            JOIN usuarios u ON os.solicitante_id = u.id
            ORDER BY os.data DESC
            LIMIT 10 
        """)
        todas_os = cursor.fetchall()
        
        stats_completas = {
            'os': stats_os_dict,
            'registros_diretos': stats_registros_dict
        }

        return render_template(
            'administrador/dashboard.html', 
            stats=stats_completas,
            todas_os=todas_os
        )

    except sqlite3.Error as e:
        flash(f'Erro ao carregar dados do dashboard: {str(e)}', 'danger')
        stats_vazias = {
            'os': {'total': 0, 'concluidas': 0, 'abertas': 0, 'em_andamento': 0, 'agendadas': 0, 'media_tempo': 0.0},
            'registros_diretos': {'total_registros': 0, 'pendente_aprovacao': 0, 'concluidos_registros': 0, 'cancelados_registros': 0}
        }
        return render_template('administrador/dashboard.html', stats=stats_vazias, todas_os=[])

@admin_bp.route('/relatorio/os')
@admin_required
def gerar_relatorio_os():
    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                os.id, os.equipamento, os.problema, os.prioridade, os.status,
                os.data as data_abertura, os.inicio as data_inicio_reparo, os.fim as data_conclusao,
                os.local, os.setor, os.solucao, os.tempo_reparo,
                u_solicitante.nome as nome_solicitante,
                u_tecnico_sistema.nome as nome_tecnico_sistema
            FROM ordens_servico os
            LEFT JOIN usuarios u_solicitante ON os.solicitante_id = u_solicitante.id
            LEFT JOIN usuarios u_tecnico_sistema ON os.tecnico_id = u_tecnico_sistema.id
            ORDER BY os.id DESC
        """)
        ordens_servico = cursor.fetchall()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Relatorio OS"

        headers = [
            "ID OS", "Equipamento", "Problema", "Prioridade", "Status", 
            "Data Abertura", "Data Início Reparo", "Data Conclusão",
            "Local", "Setor", "Solicitante", "Técnico Sistema (Agendou/Iniciou)",
            "Técnicos Participantes (Reparo)", "Solução", "Tempo Reparo (min)"
        ]
        sheet.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="004085", end_color="004085", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for col_num, header_title in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20 if len(header_title) < 15 else len(header_title) * 1.2

        for row_num, os_data_row in enumerate(ordens_servico, 2):
            os_dict = dict(os_data_row) 

            cursor.execute("""
                SELECT t.nome
                FROM participantes_os po
                JOIN tecnicos t ON po.tecnico_ref_id = t.id
                WHERE po.os_id = ?
                ORDER BY t.nome
            """, (os_dict['id'],))
            participantes_db = cursor.fetchall()
            nomes_participantes = ", ".join([p['nome'] for p in participantes_db]) if participantes_db else "N/A"

            row_data = [
                os_dict.get('id'), os_dict.get('equipamento'), os_dict.get('problema'),
                os_dict.get('prioridade'), os_dict.get('status'), os_dict.get('data_abertura'),
                os_dict.get('data_inicio_reparo', ''), os_dict.get('data_conclusao', ''),   
                os_dict.get('local', ''), os_dict.get('setor', ''), os_dict.get('nome_solicitante', ''),
                os_dict.get('nome_tecnico_sistema', ''), nomes_participantes,
                os_dict.get('solucao', ''), os_dict.get('tempo_reparo', '')
            ]
            sheet.append(row_data)
            
            for col_num in range(1, len(headers) + 1):
                sheet.cell(row=row_num, column=col_num).border = thin_border

        excel_stream = io.BytesIO()
        workbook.save(excel_stream)
        excel_stream.seek(0)

        filename = f"relatorio_os_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            excel_stream,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('admin.admin_dashboard'))

@admin_bp.route('/os/<int:id_os>')
@admin_required
def detalhe_os_admin(id_os):
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT os.*, s.nome as solicitante_nome, s.email as solicitante_email,
                   u_tecnico_sistema.nome as nome_tecnico_sistema 
            FROM ordens_servico os
            LEFT JOIN usuarios s ON os.solicitante_id = s.id
            LEFT JOIN usuarios u_tecnico_sistema ON os.tecnico_id = u_tecnico_sistema.id
            WHERE os.id = ?
        """, (id_os,))
        os_data = cursor.fetchone()
        
        if not os_data:
            flash('Ordem de Serviço não encontrada.', 'danger')
            return redirect(url_for('admin.admin_dashboard'))
        
        os_dict = dict(os_data)
        solicitante_info = {'nome': os_data['solicitante_nome'], 'email': os_data['solicitante_email']} if os_data['solicitante_nome'] else None
        tecnico_sistema_info = {'nome': os_data['nome_tecnico_sistema']} if os_data['nome_tecnico_sistema'] else None
        
        cursor.execute("""
            SELECT t.id, t.nome, t.tipo_tecnico  
            FROM participantes_os po
            JOIN tecnicos t ON po.tecnico_ref_id = t.id  
            WHERE po.os_id = ?
            ORDER BY t.nome
        """, (id_os,))
        
        participantes_individuais = [
            {'id': row['id'], 'nome': row['nome'], 'email': None, 
             'funcao': row['tipo_tecnico'].capitalize() if row['tipo_tecnico'] else 'Técnico'} 
            for row in cursor.fetchall()
        ]

        cursor.execute("""
            SELECT h.*, u.nome as usuario_nome
            FROM historico_os h
            JOIN usuarios u ON h.usuario_id = u.id
            WHERE h.os_id = ?
            ORDER BY h.data_alteracao DESC
        """, (id_os,))
        historico = cursor.fetchall()
        
        return render_template(
            'administrador/detalhes_os.html', os=os_dict, solicitante=solicitante_info,
            tecnico=tecnico_sistema_info, participantes=participantes_individuais, historico=historico
        )
        
    except Exception as e:
        flash(f'Erro ao carregar detalhes da OS: {str(e)}', 'danger')
        return redirect(url_for('admin.admin_dashboard'))

@admin_bp.route('/registros_manutencao')
@admin_required
def listar_registros_diretos():
    conn = get_db()
    cursor = conn.cursor()
    status_filtro = request.args.get('status_filtro') 
    
    query = """
        SELECT rmd.*, u_criador.nome as nome_criador
        FROM registros_manutencao_direta rmd
        JOIN usuarios u_criador ON rmd.criado_por_id = u_criador.id
    """
    params = []
    if status_filtro:
        query += " WHERE rmd.status = ?"
        params.append(status_filtro)
    query += " ORDER BY rmd.data_registro DESC"
    
    cursor.execute(query, params)
    registros = cursor.fetchall()

    return render_template('administrador/listar_registros_diretos.html', 
                           registros=registros, status_filtrado=status_filtro)

@admin_bp.route('/registros_manutencao/processar/<int:id_registro>', methods=['POST'])
@admin_required
def processar_registro_direto(id_registro):
    acao = request.form.get('acao')
    admin_id = session.get('user_id')
    data_atual_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cursor = conn.cursor()

    try:
        if acao == 'concluir':
            cursor.execute("""
                UPDATE registros_manutencao_direta
                SET status = 'Concluido', concluido_por_admin_id = ?, data_conclusao_admin = ?
                WHERE id = ? AND status = 'Pendente Aprovacao' 
            """, (admin_id, data_atual_str, id_registro))
            if cursor.rowcount > 0:
                flash('Registro de manutenção concluído com sucesso!', 'success')
            else:
                flash('Não foi possível concluir o registro.', 'warning')
        
        elif acao == 'cancelar':
            cursor.execute("""
                UPDATE registros_manutencao_direta
                SET status = 'Cancelado', concluido_por_admin_id = ?, data_conclusao_admin = ? 
                WHERE id = ? AND status = 'Pendente Aprovacao'
            """, (admin_id, data_atual_str, id_registro))
            if cursor.rowcount > 0:
                flash('Registro de manutenção cancelado.', 'info')
            else:
                flash('Não foi possível cancelar o registro.', 'warning')
        else:
            flash('Ação inválida.', 'danger')

        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao processar o registro: {str(e)}', 'danger')
    
    return redirect(url_for('admin.detalhe_registro_direto', id_registro=id_registro))

@admin_bp.route('/registros_manutencao/<int:id_registro>')
@admin_required
def detalhe_registro_direto(id_registro):
    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT rmd.*, u_criador.nome as nome_criador, u_admin.nome as nome_admin_conclusao
            FROM registros_manutencao_direta rmd
            JOIN usuarios u_criador ON rmd.criado_por_id = u_criador.id
            LEFT JOIN usuarios u_admin ON rmd.concluido_por_admin_id = u_admin.id
            WHERE rmd.id = ?
        """, (id_registro,))
        registro_data = cursor.fetchone()

        if not registro_data:
            flash('Registro de manutenção não encontrado.', 'danger')
            return redirect(url_for('admin.listar_registros_diretos'))
        
        registro_dict = dict(registro_data)

        cursor.execute("""
            SELECT t.id, t.nome, t.tipo_tecnico 
            FROM participantes_registro_direto prd
            JOIN tecnicos t ON prd.tecnico_ref_id = t.id  
            WHERE prd.registro_id = ?
            ORDER BY t.nome
        """, (id_registro,))
        
        participantes_individuais = [
            {'id': row['id'], 'nome': row['nome'], 'especialidade': row['tipo_tecnico'].capitalize() if row['tipo_tecnico'] else 'Técnico'} 
            for row in cursor.fetchall()
        ]
        
        return render_template(
            'administrador/detalhe_registro_direto.html', 
            registro=registro_dict, participantes=participantes_individuais
        )

    except Exception as e:
        flash(f'Erro ao carregar detalhes do registro: {str(e)}', 'danger')
        return redirect(url_for('admin.listar_registros_diretos'))

@admin_bp.route('/configuracoes')
@admin_required
def admin_configuracoes():
    return render_template('administrador/configuracoes.html')

@admin_bp.route('/tecnicos')
@admin_required
def listar_tecnicos():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM tecnicos ORDER BY nome")
    lista_de_tecnicos = cursor.fetchall()
    return render_template('administrador/listar_tecnicos.html', tecnicos=lista_de_tecnicos)

@admin_bp.route('/tecnicos/novo', methods=['GET', 'POST'])
@admin_required
def adicionar_tecnico():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo_tecnico = request.form.get('tipo_tecnico')

        if not nome or not tipo_tecnico:
            flash('Nome e Tipo do Técnico são obrigatórios.', 'warning')
            return render_template('administrador/form_tecnico.html', tecnico=request.form, titulo="Adicionar Novo Técnico", acao="Adicionar")
        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO tecnicos (nome, tipo_tecnico, ativo) VALUES (?, ?, 1)", (nome, tipo_tecnico))
            conn.commit()
            flash('Técnico adicionado com sucesso!', 'success')
            return redirect(url_for('admin.listar_tecnicos'))
        except sqlite3.IntegrityError:
            flash('Já existe um técnico com este nome.', 'danger')
        except Exception as e:
            flash(f'Erro ao adicionar técnico: {str(e)}', 'danger')
        
        return render_template('administrador/form_tecnico.html', tecnico=request.form, titulo="Adicionar Novo Técnico", acao="Adicionar")
        
    return render_template('administrador/form_tecnico.html', titulo="Adicionar Novo Técnico", acao="Adicionar", tecnico=None)

@admin_bp.route('/tecnicos/editar/<int:id_tecnico>', methods=['GET', 'POST'])
@admin_required
def editar_tecnico(id_tecnico):
    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        tipo_tecnico = request.form.get('tipo_tecnico')
        ativo = 1 if request.form.get('ativo') == '1' else 0 

        if not nome or not tipo_tecnico:
            flash('Nome e Tipo do Técnico são obrigatórios.', 'warning')
            cursor.execute("SELECT * FROM tecnicos WHERE id = ?", (id_tecnico,))
            tecnico_data = cursor.fetchone()
            if not tecnico_data:
                flash('Técnico não encontrado.', 'danger')
                return redirect(url_for('admin.listar_tecnicos'))
            
            form_data_com_erro = {'id': id_tecnico, 'nome': nome, 'tipo_tecnico': tipo_tecnico, 'ativo': ativo}
            return render_template('administrador/form_tecnico.html', tecnico=form_data_com_erro, titulo=f"Editar Técnico: {tecnico_data['nome']}", acao="Editar")

        try:
            cursor.execute("SELECT id FROM tecnicos WHERE nome = ? AND id != ?", (nome, id_tecnico))
            if cursor.fetchone():
                flash('Já existe outro técnico com este nome.', 'danger')
            else:
                cursor.execute("UPDATE tecnicos SET nome = ?, tipo_tecnico = ?, ativo = ? WHERE id = ?", (nome, tipo_tecnico, ativo, id_tecnico))
                conn.commit()
                flash('Técnico atualizado com sucesso!', 'success')
                return redirect(url_for('admin.listar_tecnicos'))
        except Exception as e:
            flash(f'Erro ao atualizar técnico: {str(e)}', 'danger')

        cursor.execute("SELECT nome FROM tecnicos WHERE id = ?", (id_tecnico,))
        nome_original = cursor.fetchone()
        nome_original_tecnico = nome_original['nome'] if nome_original else f"ID {id_tecnico}"
        form_data_com_erro = {'id': id_tecnico, 'nome': nome, 'tipo_tecnico': tipo_tecnico, 'ativo': ativo}
        return render_template('administrador/form_tecnico.html', tecnico=form_data_com_erro, titulo=f"Editar Técnico: {nome_original_tecnico}", acao="Editar")

    cursor.execute("SELECT * FROM tecnicos WHERE id = ?", (id_tecnico,))
    tecnico_data = cursor.fetchone()

    if not tecnico_data:
        flash('Técnico não encontrado.', 'danger')
        return redirect(url_for('admin.listar_tecnicos'))

    return render_template('administrador/form_tecnico.html', tecnico=tecnico_data, titulo=f"Editar Técnico: {tecnico_data['nome']}", acao="Editar")

@admin_bp.route('/locais')
@admin_required
def listar_locais():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM locais ORDER BY nome")
    locais = cursor.fetchall()
    return render_template('administrador/listar_locais.html', locais=locais)

@admin_bp.route('/locais/adicionar', methods=['GET', 'POST'])
@admin_required
def adicionar_local():
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        ativo = 1 if request.form.get('ativo') == '1' else 0

        if not nome:
            flash('O nome do local é obrigatório.', 'danger')
            return render_template('administrador/form_local.html', local=None)

        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute("INSERT INTO locais (nome, descricao, ativo) VALUES (?, ?, ?)", (nome, descricao, ativo))
            conn.commit()
            flash('Local adicionado com sucesso!', 'success')
            return redirect(url_for('admin.listar_locais'))
        except sqlite3.IntegrityError:
            conn.rollback()
            flash('Já existe um local com este nome.', 'danger')
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao adicionar local: {str(e)}', 'danger')
            
        return render_template('administrador/form_local.html', local={'nome': nome, 'descricao': descricao, 'ativo': ativo})

    return render_template('administrador/form_local.html', local=None)

@admin_bp.route('/locais/editar/<int:id>', methods=['GET', 'POST'])
@admin_required
def editar_local(id):
    conn = get_db() 
    cursor = conn.cursor()

    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        descricao = request.form.get('descricao', '').strip()
        ativo = 1 if request.form.get('ativo') == '1' else 0

        if not nome:
            flash('O nome do local é obrigatório.', 'danger')
            cursor.execute("SELECT * FROM locais WHERE id = ?", (id,))
            local_data = cursor.fetchone()
            if not local_data:
                flash('Local não encontrado.', 'danger')
                return redirect(url_for('admin.listar_locais'))
            local_dict_for_template = dict(local_data)
            local_dict_for_template.update({'nome': nome, 'descricao': descricao, 'ativo': ativo})
            return render_template('administrador/form_local.html', local=local_dict_for_template)

        try:
            cursor.execute("SELECT id FROM locais WHERE nome = ? AND id != ?", (nome, id))
            if cursor.fetchone():
                flash('Já existe outro local com este nome.', 'danger')
            else:
                cursor.execute("UPDATE locais SET nome = ?, descricao = ?, ativo = ? WHERE id = ?", (nome, descricao, ativo, id))
                conn.commit()
                flash('Local atualizado com sucesso!', 'success')
                return redirect(url_for('admin.listar_locais'))
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao atualizar local: {str(e)}', 'danger')

    cursor.execute("SELECT * FROM locais WHERE id = ?", (id,))
    local_data = cursor.fetchone()

    if not local_data:
        flash('Local não encontrado.', 'danger')
        return redirect(url_for('admin.listar_locais'))
    
    return render_template('administrador/form_local.html', local=local_data)

@admin_bp.route('/locais/remover/<int:id>')
@admin_required
def remover_local(id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM locais WHERE id = ?", (id,))
        conn.commit()
        flash('Local removido com sucesso!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao remover local: {str(e)}', 'danger')
    
    return redirect(url_for('admin.listar_locais'))

@admin_bp.route('/usuarios')
@admin_required
def listar_usuarios():
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios ORDER BY nome")
    usuarios = cursor.fetchall()
    return render_template('administrador/listar_usuarios.html', usuarios=usuarios)

@admin_bp.route('/usuarios/novo', methods=['GET', 'POST'])
@admin_required
def novo_usuario():
    tipos_disponiveis_para_criacao = []
    if session.get('tipo') == 'master-admin':
        tipos_disponiveis_para_criacao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Login da Manutenção (Compartilhado)'},
            {'value': 'admin', 'label': 'Administrador'}
        ]
    elif session.get('tipo') == 'admin':
        tipos_disponiveis_para_criacao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Login da Manutenção (Compartilhado)'} 
        ]

    if request.method == 'POST':
        usuario_form = request.form.get('usuario', '').strip()
        senha_form = request.form.get('senha')
        tipo_novo_usuario = request.form.get('tipo')
        nome_form = request.form.get('nome', '').strip()
        email_form = request.form.get('email', '').strip()

        if not all([usuario_form, senha_form, tipo_novo_usuario, nome_form, email_form]):
            flash('Todos os campos marcados com * são obrigatórios.', 'warning')
            return render_template('administrador/novo_usuario.html', tipos_disponiveis=tipos_disponiveis_para_criacao, submitted_data=request.form)
        
        tipos_valores_permitidos = [t['value'] for t in tipos_disponiveis_para_criacao]
        if tipo_novo_usuario not in tipos_valores_permitidos:
            flash('Você não tem permissão para criar este tipo de usuário.', 'danger')
            return render_template('administrador/novo_usuario.html', tipos_disponiveis=tipos_disponiveis_para_criacao, submitted_data=request.form)

        try:
            conn = get_db()
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO usuarios (usuario, senha, tipo, nome, email, ativo)
                VALUES (?, ?, ?, ?, ?, 1)
            ''', (usuario_form, generate_password_hash(senha_form), tipo_novo_usuario, nome_form, email_form))
            conn.commit()
            flash('Usuário criado com sucesso!', 'success')
            return redirect(url_for('admin.listar_usuarios'))
        except sqlite3.IntegrityError:
            flash('Nome de usuário ou email já cadastrado.', 'danger')
        except Exception as e:
            flash(f'Erro ao criar usuário: {str(e)}', 'danger')
        
        return render_template('administrador/novo_usuario.html', tipos_disponiveis=tipos_disponiveis_para_criacao, submitted_data=request.form)

    return render_template('administrador/novo_usuario.html', tipos_disponiveis=tipos_disponiveis_para_criacao, submitted_data=None)

@admin_bp.route('/usuarios/editar/<int:id_usuario_alvo>', methods=['GET', 'POST'])
@admin_required
def editar_usuario(id_usuario_alvo):
    editor_tipo = session.get('tipo')
    editor_id = session.get('user_id')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM usuarios WHERE id = ?", (id_usuario_alvo,))
    usuario_alvo = cursor.fetchone()

    if not usuario_alvo:
        flash('Usuário não encontrado!', 'danger')
        return redirect(url_for('admin.listar_usuarios'))

    pode_editar = False
    tipos_disponiveis_para_atribuicao = []

    if editor_tipo == 'master-admin':
        pode_editar = True
        tipos_disponiveis_para_atribuicao = [
            {'value': 'solicitante', 'label': 'Solicitante'},
            {'value': 'manutencao', 'label': 'Técnico de Manutenção'},
            {'value': 'admin', 'label': 'Administrador'}
        ]
        if usuario_alvo['tipo'] == 'master-admin':
            tipos_disponiveis_para_atribuicao = [{'value': 'master-admin', 'label': 'Master Administrador'}] 
            if usuario_alvo['id'] == editor_id and sum(1 for row in cursor.execute("SELECT 1 FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1")) <= 1:
                 tipos_disponiveis_para_atribuicao = [{'value': 'master-admin', 'label': 'Master Administrador'}]

    elif editor_tipo == 'admin':
        if usuario_alvo['tipo'] == 'master-admin':
            flash('Administradores não podem editar usuários Master Administrador.', 'warning')
        elif usuario_alvo['tipo'] == 'admin' and usuario_alvo['id'] != editor_id:
            flash('Administradores não podem editar outros Administradores.', 'warning')
        elif usuario_alvo['id'] == editor_id: 
            pode_editar = True 
            tipos_disponiveis_para_atribuicao = [{'value': 'admin', 'label': 'Administrador'}] 
        else: 
            pode_editar = True
            tipos_disponiveis_para_atribuicao = [
                {'value': 'solicitante', 'label': 'Solicitante'},
                {'value': 'manutencao', 'label': 'Técnico de Manutenção'}
            ]
            if usuario_alvo['tipo'] == 'admin':
                 tipos_disponiveis_para_atribuicao = [{'value': 'admin', 'label': 'Administrador'}]

    if not pode_editar and request.method == 'GET': 
         return redirect(url_for('admin.listar_usuarios'))

    if request.method == 'POST':
        if not pode_editar: 
            flash('Você não tem permissão para modificar este usuário.', 'danger')
            return redirect(url_for('admin.listar_usuarios'))

        nome_form = request.form.get('nome','').strip()
        usuario_form = request.form.get('usuario','').strip()
        email_form = request.form.get('email','').strip()
        tipo_form = request.form.get('tipo')
        ativo_form = 1 if request.form.get('ativo') == 'on' else 0 
        especialidade_form = request.form.get('especialidade') if tipo_form == 'manutencao' else None

        if not all([nome_form, usuario_form, email_form, tipo_form]):
            flash('Campos Nome, Usuário, Email e Tipo são obrigatórios.', 'warning')
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', usuario=usuario_alvo_temp_form, tipos_disponiveis=tipos_disponiveis_para_atribuicao, pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

        tipos_valores_permitidos = [t['value'] for t in tipos_disponiveis_para_atribuicao]
        if tipo_form not in tipos_valores_permitidos:
            flash('Você não tem permissão para atribuir este tipo de usuário.', 'danger')
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', usuario=usuario_alvo_temp_form, tipos_disponiveis=tipos_disponiveis_para_atribuicao, pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

        if tipo_form == 'manutencao' and not especialidade_form:
            flash('Especialidade é obrigatória para Técnico de Manutenção.', 'warning')
            usuario_alvo_temp_form = dict(usuario_alvo)
            usuario_alvo_temp_form.update(request.form.to_dict())
            return render_template('administrador/editar_usuario.html', usuario=usuario_alvo_temp_form, tipos_disponiveis=tipos_disponiveis_para_atribuicao, pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))
        
        if usuario_alvo['tipo'] == 'master-admin' and usuario_alvo['id'] == editor_id:
            if not ativo_form:
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1 AND id != ?", (editor_id,))
                if cursor.fetchone()[0] == 0:
                    flash('Não é possível desativar o único Master Administrador ativo.', 'danger')
                    ativo_form = 1 
            if tipo_form != 'master-admin':
                 flash('Master Administradores não podem mudar seu próprio tipo desta forma.', 'danger')
                 tipo_form = 'master-admin' 

        try:
            cursor.execute('''
                UPDATE usuarios 
                SET nome = ?, usuario = ?, email = ?, tipo = ?, especialidade = ?, ativo = ?
                WHERE id = ?
            ''', (nome_form, usuario_form, email_form, tipo_form, especialidade_form, ativo_form, id_usuario_alvo))
            conn.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('admin.listar_usuarios'))
        except sqlite3.IntegrityError:
            flash('Nome de usuário ou email já existe para outro usuário.', 'danger')
        except Exception as e:
            flash(f'Erro ao atualizar usuário: {str(e)}', 'danger')
        
        usuario_alvo_temp_form = dict(usuario_alvo)
        usuario_alvo_temp_form.update(request.form.to_dict()) 
        return render_template('administrador/editar_usuario.html', usuario=usuario_alvo_temp_form, tipos_disponiveis=tipos_disponiveis_para_atribuicao, pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

    return render_template('administrador/editar_usuario.html', usuario=usuario_alvo, tipos_disponiveis=tipos_disponiveis_para_atribuicao, pode_alterar_tipo=(usuario_alvo['tipo'] != 'master-admin' and (editor_tipo == 'master-admin' or usuario_alvo['id'] != editor_id)))

@admin_bp.route('/usuarios/alterar_senha/<int:id_usuario_alvo>', methods=['POST'])
@admin_required
def alterar_senha_usuario(id_usuario_alvo):
    nova_senha = request.form.get('nova_senha')
    if not nova_senha:
        flash('Nova senha não pode ser vazia.', 'warning')
        return redirect(url_for('admin.editar_usuario', id_usuario_alvo=id_usuario_alvo)) 
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE usuarios 
            SET senha = ?
            WHERE id = ?
        ''', (generate_password_hash(nova_senha), id_usuario_alvo))
        conn.commit()
        flash('Senha alterada com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao alterar senha: {str(e)}', 'danger')
    
    return redirect(url_for('admin.editar_usuario', id_usuario_alvo=id_usuario_alvo))

@admin_bp.route('/usuarios/remover/<int:id_usuario_alvo>')
@admin_required
def remover_usuario(id_usuario_alvo):
    editor_tipo = session.get('tipo')
    editor_id = session.get('user_id')

    if id_usuario_alvo == editor_id:
        flash('Você não pode remover a si mesmo!', 'danger')
        return redirect(url_for('admin.listar_usuarios'))

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM usuarios WHERE id = ?", (id_usuario_alvo,))
    usuario_alvo = cursor.fetchone()

    if not usuario_alvo:
        flash('Usuário a ser removido não encontrado.', 'warning')
        return redirect(url_for('admin.listar_usuarios'))

    pode_remover = False
    mensagem_erro_permissao = 'Você não tem permissão para remover este tipo de usuário.'

    if editor_tipo == 'master-admin':
        if usuario_alvo['tipo'] == 'master-admin':
            if usuario_alvo['id'] == editor_id: 
                pode_remover = False
                mensagem_erro_permissao = 'Você não pode remover a si mesmo!'
            else:
                cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1 AND id != ?", (id_usuario_alvo,))
                if cursor.fetchone()[0] >= 1:
                    pode_remover = True
                else:
                    cursor.execute("SELECT COUNT(*) FROM usuarios WHERE tipo = 'master-admin' AND ativo = 1")
                    if cursor.fetchone()[0] > 1:
                        pode_remover = True
                    else:
                        pode_remover = False
                        mensagem_erro_permissao = 'Não é possível remover o único Master Administrador ativo.'
        else:
            pode_remover = True
    
    elif editor_tipo == 'admin':
        if usuario_alvo['tipo'] == 'master-admin':
            pode_remover = False
            mensagem_erro_permissao = 'Administradores não podem remover usuários Master Administrador.'
        elif usuario_alvo['tipo'] == 'admin':
            pode_remover = False
            mensagem_erro_permissao = 'Administradores não podem remover outros Administradores.'
        elif usuario_alvo['tipo'] in ['solicitante', 'manutencao']:
            pode_remover = True

    if not pode_remover:
        flash(mensagem_erro_permissao, 'danger')
        return redirect(url_for('admin.listar_usuarios'))

    try:
        cursor.execute("SELECT COUNT(*) FROM ordens_servico WHERE solicitante_id = ? OR tecnico_id = ?", (id_usuario_alvo, id_usuario_alvo))
        if cursor.fetchone()[0] > 0:
            flash(f"Não é possível remover o usuário (ID: {id_usuario_alvo}) pois ele está associado a Ordens de Serviço. Considere desativá-lo.", 'warning')
            return redirect(url_for('admin.listar_usuarios'))

        cursor.execute("SELECT COUNT(*) FROM registros_manutencao_direta WHERE criado_por_id = ? OR concluido_por_admin_id = ?", (id_usuario_alvo, id_usuario_alvo))
        if cursor.fetchone()[0] > 0:
            flash(f"Não é possível remover o usuário (ID: {id_usuario_alvo}) pois ele está associado a Registros de Manutenção. Considere desativá-lo.", 'warning')
            return redirect(url_for('admin.listar_usuarios'))

        cursor.execute("DELETE FROM participantes_os WHERE tecnico_id = ?", (id_usuario_alvo,))
        cursor.execute("DELETE FROM participantes_registro_direto WHERE tecnico_id = ?", (id_usuario_alvo,))
        cursor.execute("DELETE FROM usuarios WHERE id = ?", (id_usuario_alvo,))
        conn.commit()
        flash('Usuário removido com sucesso!', 'success')
    except sqlite3.IntegrityError as e:
        flash(f"Erro de integridade ao remover usuário. Verifique se o usuário está referenciado em outras tabelas.", 'danger')
    except Exception as e:
        flash(f'Erro ao remover usuário: {str(e)}', 'danger')
    
    return redirect(url_for('admin.listar_usuarios'))